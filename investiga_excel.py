#!/usr/bin/env python3
"""
Script exhaustivo para investigar el archivo Excel DISTRIKIA COPIA.xlsx
Busca fórmulas, tablas dinámicas, named ranges, datos ocultos, comentarios, gráficos, etc.
Salida guardada en investigacion_excel.txt
"""

import sys
import json
from collections import Counter
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd

output_file = "investigacion_excel.txt"
f = open(output_file, "w", encoding="utf-8")

def out(*args, **kwargs):
    print(*args, **kwargs)
    print(*args, **kwargs, file=f)

def main():
    file_path = "DISTRIKIA COPIA.xlsx"
    out("=" * 80)
    out(f"INVESTIGACIÓN EXHAUSTIVA: {file_path}")
    out("=" * 80)

    # Cargar el workbook
    wb = load_workbook(file_path, data_only=False)
    wb_data = load_workbook(file_path, data_only=True)

    # ============================================================
    # 1. PROPIEDADES DEL ARCHIVO / METADATA
    # ============================================================
    out("\n" + "=" * 80)
    out("1. PROPIEDADES DEL ARCHIVO / METADATA")
    out("=" * 80)
    attrs = ['creator', 'lastModifiedBy', 'created', 'modified', 'title', 'subject', 'category', 'keywords', 'description', 'version', 'contentStatus', 'identifier', 'language', 'lastPrinted', 'revision']
    for attr in attrs:
        try:
            val = getattr(wb.properties, attr, None)
            out(f"  {attr}: {val}")
        except Exception as e:
            out(f"  {attr}: ERROR - {e}")

    # ============================================================
    # 2. HOJAS DISPONIBLES
    # ============================================================
    out("\n" + "=" * 80)
    out("2. HOJAS DISPONIBLES EN EL LIBRO")
    out("=" * 80)
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        out(f"  {i}. '{sheet_name}' - dimensiones: {ws.dimensions}, filas usadas: {ws.max_row}, cols usadas: {ws.max_column}")

    # ============================================================
    # 3. NAMED RANGES
    # ============================================================
    out("\n" + "=" * 80)
    out("3. NAMED RANGES (RANGOS DEFINIDOS)")
    out("=" * 80)
    if wb.defined_names:
        for name, defined_name in wb.defined_names.items():
            out(f"  Name: '{name}' -> {defined_name.attr_text}")
    else:
        out("  No hay named ranges definidos.")

    # ============================================================
    # 4. ANÁLISIS POR HOJA - RESUMEN
    # ============================================================
    all_sheets = wb.sheetnames

    for sheet_name in all_sheets:
        out("\n" + "=" * 80)
        out(f"4. ANÁLISIS DETALLADO DE HOJA: '{sheet_name}'")
        out("=" * 80)

        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name]

        out(f"  Dimensiones: {ws.dimensions}")
        out(f"  Max fila: {ws.max_row}, Max columna: {ws.max_column}")
        out(f"  Sheet state: {ws.sheet_state}")
        out(f"  Sheet properties: tabColor={ws.sheet_properties.tabColor}")

        hidden_rows = [r for r in range(1, ws.max_row + 1) if ws.row_dimensions[r].hidden]
        hidden_cols = [get_column_letter(c) for c in range(1, ws.max_column + 1) if ws.column_dimensions[get_column_letter(c)].hidden]
        out(f"  Filas ocultas: {len(hidden_rows)} {'- ' + str(hidden_rows[:20]) if hidden_rows else '(ninguna)'}")
        out(f"  Columnas ocultas: {len(hidden_cols)} {'- ' + str(hidden_cols) if hidden_cols else '(ninguna)'}")
        out(f"  Hoja protegida: {ws.protection.sheet}")

        if ws.tables:
            out(f"  Tablas Excel: {list(ws.tables.keys())}")
            for table_name in ws.tables:
                table = ws.tables[table_name]
                col_names = [c.name for c in table.tableColumns]
                out(f"    Tabla '{table_name}': rango={table.ref}, cols={col_names}")
        else:
            out("  No hay tablas Excel.")

        if ws._charts:
            out(f"  Gráficos incrustados: {len(ws._charts)}")
            for idx, chart in enumerate(ws._charts, 1):
                out(f"    Gráfico {idx}: tipo={chart.__class__.__name__}, title={chart.title}")
        else:
            out("  No hay gráficos incrustados.")

        comments_found = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    comments_found.append((cell.coordinate, cell.comment.text))
        out(f"  Comentarios: {len(comments_found)}")
        for coord, text in comments_found[:10]:
            out(f"    {coord}: {text[:200]}")

        # FÓRMULAS - resumen por patrón
        formula_patterns = Counter()
        formula_cells = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_patterns[cell.value] += 1
                    formula_cells.append((cell.coordinate, cell.value))

        out(f"  Total fórmulas: {len(formula_cells)}")
        if formula_patterns:
            out(f"  Patrones de fórmula únicos: {len(formula_patterns)}")
            for formula, count in formula_patterns.most_common(20):
                out(f"    ({count}x) {formula}")

        # Buscar celdas con palabras clave
        keywords = ['tasa', 'imprevisto', 'imprevistos', 'porcentaje', '%', 'recuperación', 'recuperacion', 'base', 'datos']
        keyword_cells = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    lower_val = cell.value.lower()
                    for kw in keywords:
                        if kw in lower_val:
                            keyword_cells.append((cell.coordinate, cell.value))
                            break
        out(f"  Celdas con palabras clave: {len(keyword_cells)}")
        for coord, val in keyword_cells[:30]:
            out(f"    {coord}: '{val}'")

        # Buscar valores numéricos que parezcan porcentajes
        percentage_like = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    v = cell.value
                    if (0.20 <= v <= 0.70) or (20 <= v <= 70):
                        percentage_like.append((cell.coordinate, v))
        out(f"  Valores tipo porcentaje (0.20-0.70 o 20-70): {len(percentage_like)}")
        for coord, val in percentage_like[:40]:
            out(f"    {coord}: {val}")

        # Mostrar contenido completo si la hoja es pequeña
        total_cells = ws.max_row * ws.max_column
        if total_cells <= 500:
            out(f"  CONTENIDO COMPLETO:")
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                row_data = []
                for cell in row:
                    if cell.value is not None:
                        row_data.append(f"{cell.coordinate}={repr(cell.value)}")
                if row_data:
                    out(f"    {' | '.join(row_data)}")
        else:
            out(f"  Hoja grande ({total_cells} celdas), contenido completo omitido.")

    # ============================================================
    # 5. BÚSQUEDA GLOBAL DE PORCENTAJES DEL CLIENTE
    # ============================================================
    out("\n" + "=" * 80)
    out("5. BÚSQUEDA GLOBAL DE LOS PORCENTAJES DEL CLIENTE")
    out("=" * 80)
    client_pcts = [0.40, 0.47, 0.60, 0.42, 0.33, 0.44, 0.41, 0.50, 0.44, 0.35, 0.24, 0.44, 0.54, 0.60, 0.42]
    client_pcts_int = [40, 47, 60, 42, 33, 44, 41, 50, 44, 35, 24, 44, 54, 60, 42]

    for sheet_name in all_sheets:
        ws = wb[sheet_name]
        matches = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    v = cell.value
                    if v in client_pcts or v in client_pcts_int or round(v, 2) in [round(p, 2) for p in client_pcts]:
                        matches.append((cell.coordinate, v))
        if matches:
            out(f"  Hoja '{sheet_name}' - coincidencias: {len(matches)}")
            for coord, val in matches[:30]:
                out(f"    {sheet_name}!{coord}: {val}")

    # ============================================================
    # 6. VBA / MACROS
    # ============================================================
    out("\n" + "=" * 80)
    out("6. VBA / MACROS")
    out("=" * 80)
    try:
        if wb.vba_archive:
            out(f"  ⚠️  El archivo CONTIENE código VBA/macros.")
            out(f"  VBA archive contents: {wb.vba_archive.namelist()[:20]}")
        else:
            out("  No se detectó código VBA.")
    except Exception as e:
        out(f"  Error revisando VBA: {e}")

    # ============================================================
    # 7. PROTECCIÓN DEL LIBRO
    # ============================================================
    out("\n" + "=" * 80)
    out("7. PROTECCIÓN DEL LIBRO")
    out("=" * 80)
    try:
        if wb.security.workbookPassword:
            out("  ⚠️  Libro protegido con contraseña.")
        else:
            out("  Libro no protegido con contraseña.")
    except Exception as e:
        out(f"  Error: {e}")

    # ============================================================
    # 8. DATA VALIDATION
    # ============================================================
    out("\n" + "=" * 80)
    out("8. DATA VALIDATION")
    out("=" * 80)
    for sheet_name in all_sheets:
        ws = wb[sheet_name]
        dv_list = list(ws.data_validations.dataValidation)
        if dv_list:
            out(f"  Hoja '{sheet_name}': {len(dv_list)} reglas de validación")
            for dv in dv_list[:5]:
                out(f"    Rango: {dv.sqref}, Tipo: {dv.type}, Fórmula: {dv.formula1}")
        else:
            out(f"  Hoja '{sheet_name}': sin validación de datos.")

    # ============================================================
    # 9. HYPERLINKS
    # ============================================================
    out("\n" + "=" * 80)
    out("9. HYPERLINKS")
    out("=" * 80)
    for sheet_name in all_sheets:
        ws = wb[sheet_name]
        links = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    links.append((cell.coordinate, str(cell.hyperlink.target) if cell.hyperlink.target else str(cell.hyperlink)))
        if links:
            out(f"  Hoja '{sheet_name}': {len(links)} hyperlinks")
            for coord, target in links[:10]:
                out(f"    {coord}: {target}")
        else:
            out(f"  Hoja '{sheet_name}': sin hyperlinks.")

    # ============================================================
    # 10. PRINT AREA / PRINT TITLES
    # ============================================================
    out("\n" + "=" * 80)
    out("10. PRINT AREA / PRINT TITLES")
    out("=" * 80)
    for sheet_name in all_sheets:
        ws = wb[sheet_name]
        if ws.print_area:
            out(f"  Hoja '{sheet_name}': print_area = {ws.print_area}")
        else:
            out(f"  Hoja '{sheet_name}': sin print_area definida.")

    # ============================================================
    # 11. MERGED CELLS
    # ============================================================
    out("\n" + "=" * 80)
    out("11. MERGED CELLS")
    out("=" * 80)
    for sheet_name in all_sheets:
        ws = wb[sheet_name]
        if ws.merged_cells.ranges:
            out(f"  Hoja '{sheet_name}': {len(ws.merged_cells.ranges)} rangos combinados")
            for mr in list(ws.merged_cells.ranges)[:10]:
                out(f"    {mr}")
        else:
            out(f"  Hoja '{sheet_name}': sin celdas combinadas.")

    # ============================================================
    # 12. AUTO FILTERS
    # ============================================================
    out("\n" + "=" * 80)
    out("12. AUTO FILTERS")
    out("=" * 80)
    for sheet_name in all_sheets:
        ws = wb[sheet_name]
        if ws.auto_filter:
            out(f"  Hoja '{sheet_name}': auto_filter = {ws.auto_filter.ref}")
        else:
            out(f"  Hoja '{sheet_name}': sin auto_filter.")

    # ============================================================
    # 13. FORMATO CONDICIONAL
    # ============================================================
    out("\n" + "=" * 80)
    out("13. FORMATO CONDICIONAL")
    out("=" * 80)
    for sheet_name in all_sheets:
        ws = wb[sheet_name]
        if ws.conditional_formatting:
            out(f"  Hoja '{sheet_name}': {len(ws.conditional_formatting)} reglas")
            for cf in ws.conditional_formatting:
                out(f"    Rango: {cf}, Reglas: {len(cf.rules)}")
        else:
            out(f"  Hoja '{sheet_name}': sin formato condicional.")

    # ============================================================
    # 14. OUTLINE / GROUPING
    # ============================================================
    out("\n" + "=" * 80)
    out("14. OUTLINE / GROUPING")
    out("=" * 80)
    for sheet_name in all_sheets:
        ws = wb[sheet_name]
        row_groups = [(r, ws.row_dimensions[r].outline_level) for r in range(1, ws.max_row + 1) if ws.row_dimensions[r].outline_level and ws.row_dimensions[r].outline_level > 0]
        col_groups = [(get_column_letter(c), ws.column_dimensions[get_column_letter(c)].outline_level) for c in range(1, ws.max_column + 1) if ws.column_dimensions[get_column_letter(c)].outline_level and ws.column_dimensions[get_column_letter(c)].outline_level > 0]
        if row_groups or col_groups:
            out(f"  Hoja '{sheet_name}': filas agrupadas={len(row_groups)}, cols agrupadas={len(col_groups)}")
        else:
            out(f"  Hoja '{sheet_name}': sin grouping.")

    # ============================================================
    # 15. CONTENIDO COMPLETO DE HOJAS CLAVE
    # ============================================================
    out("\n" + "=" * 80)
    out("15. CONTENIDO COMPLETO DE HOJAS CLAVE")
    out("=" * 80)
    for target in ["TASA DE IMPREVISTOS", "RECUPERACIÓN", "CAUSALES", "ESTATUS", "REVISADOS"]:
        if target in wb.sheetnames:
            ws = wb[target]
            out(f"\n--- HOJA: '{target}' (CONTENIDO COMPLETO) ---")
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                row_data = []
                for cell in row:
                    if cell.value is not None:
                        row_data.append(f"{cell.coordinate}={repr(cell.value)}")
                if row_data:
                    out(f"  {' | '.join(row_data)}")

    out("\n" + "=" * 80)
    out("INVESTIGACIÓN COMPLETADA")
    out("=" * 80)
    f.close()
    print(f"\nResultados guardados en: {output_file}")

if __name__ == "__main__":
    main()
