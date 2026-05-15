#!/usr/bin/env python3
"""
Análisis profundo de DISTRIKIA COPIA.xlsx enfocado en la tasa de imprevistos.
"""
import pandas as pd
from openpyxl import load_workbook
import json

file_path = "DISTRIKIA COPIA.xlsx"

# Porcentajes del cliente
client_pcts = {
    'FEB': 40, 'MAR': 47, 'ABR': 60, 'MAY': 42, 'JUN': 33, 'JUL': 44,
    'AGOS': 41, 'SEPT': 50, 'OCT': 44, 'NOV': 35, 'DIC': 24, 'ENE': 44,
    'FEB2': 54, 'MAR2': 60, 'ABR2': 42
}

print("=" * 80)
print("ANÁLISIS PROFUNDO - TASA DE IMPREVISTOS")
print("=" * 80)

# Leer todas las hojas con pandas
xls = pd.ExcelFile(file_path)
print("\nHojas disponibles:", xls.sheet_names)

# ============================================================
# 1. Revisar TASA DE IMPREVISTOS
# ============================================================
print("\n" + "=" * 80)
print("1. HOJA 'TASA DE IMPREVISTOS'")
print("=" * 80)
df_tasa = pd.read_excel(file_path, sheet_name='TASA DE IMPREVISTOS', header=None)
print(df_tasa.to_string())

# ============================================================
# 2. Revisar CAUSALES en detalle
# ============================================================
print("\n" + "=" * 80)
print("2. HOJA 'CAUSALES' COMPLETA")
print("=" * 80)
df_causales = pd.read_excel(file_path, sheet_name='CAUSALES', header=None)
print(df_causales.to_string())

# ============================================================
# 3. Revisar BASE DE DATOS - primeras y últimas filas
# ============================================================
print("\n" + "=" * 80)
print("3. HOJA 'BASE DE DATOS' - ENCABEZADOS Y MUESTRA")
print("=" * 80)
df_base = pd.read_excel(file_path, sheet_name='BASE DE DATOS')
print("Columnas:", df_base.columns.tolist())
print("\nPrimeras 5 filas:")
print(df_base.head())
print("\nÚltimas 5 filas:")
print(df_base.tail())

# ============================================================
# 4. Contar registros por mes en BASE DE DATOS
# ============================================================
print("\n" + "=" * 80)
print("4. CONTEO POR MES EN BASE DE DATOS")
print("=" * 80)
if 'MES' in df_base.columns:
    print(df_base['MES'].value_counts().sort_index())
if 'FECHA INGR' in df_base.columns:
    df_base['FECHA INGR'] = pd.to_datetime(df_base['FECHA INGR'], errors='coerce')
    df_base['MES_NOMBRE'] = df_base['FECHA INGR'].dt.strftime('%b').str.upper()
    # Mapear nombres
    mes_map = {'FEB': 'FEB', 'MAR': 'MAR', 'APR': 'ABR', 'MAY': 'MAY', 'JUN': 'JUN', 'JUL': 'JUL',
               'AUG': 'AGOS', 'SEP': 'SEPT', 'OCT': 'OCT', 'NOV': 'NOV', 'DEC': 'DIC', 'JAN': 'ENE'}
    df_base['MES_NOMBRE_ES'] = df_base['MES_NOMBRE'].map(mes_map).fillna(df_base['MES_NOMBRE'])
    print("\nConteo por mes (nombre español):")
    print(df_base['MES_NOMBRE_ES'].value_counts())

# ============================================================
# 5. Buscar los porcentajes del cliente en TODO el archivo
# ============================================================
print("\n" + "=" * 80)
print("5. BÚSQUEDA EXACTA DE PORCENTAJES DEL CLIENTE EN TODAS LAS HOJAS")
print("=" * 80)

target_values = [0.40, 0.47, 0.60, 0.42, 0.33, 0.44, 0.41, 0.50, 0.35, 0.24, 0.54,
                 40, 47, 60, 42, 33, 44, 41, 50, 35, 24, 54]

for sheet in xls.sheet_names:
    df = pd.read_excel(file_path, sheet_name=sheet, header=None)
    found = []
    for col in df.columns:
        for idx, val in df[col].items():
            if pd.notna(val) and isinstance(val, (int, float)):
                if val in target_values or (isinstance(val, float) and round(val, 2) in [round(v, 2) for v in target_values]):
                    found.append((idx, col, val))
    if found:
        print(f"\nHoja '{sheet}': {len(found)} coincidencias")
        for r, c, v in found[:20]:
            print(f"  Fila {r}, Col {c}: {v}")

# ============================================================
# 6. Revisar si hay fórmulas en TASA DE IMPREVISTOS con openpyxl
# ============================================================
print("\n" + "=" * 80)
print("6. FÓRMULAS EN 'TASA DE IMPREVISTOS' (openpyxl)")
print("=" * 80)
wb = load_workbook(file_path, data_only=False)
ws = wb['TASA DE IMPREVISTOS']
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"  {cell.coordinate}: {cell.value}")

# ============================================================
# 7. Revisar RECUPERACIÓN
# ============================================================
print("\n" + "=" * 80)
print("7. HOJA 'RECUPERACIÓN' COMPLETA")
print("=" * 80)
df_rec = pd.read_excel(file_path, sheet_name='RECUPERACIÓN', header=None)
print(df_rec.to_string())

# ============================================================
# 8. Revisar si hay más hojas ocultas o con nombres similares
# ============================================================
print("\n" + "=" * 80)
print("8. NOMBRES DE HOJAS Y ESTADOS")
print("=" * 80)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  '{name}': state={ws.sheet_state}, max_row={ws.max_row}, max_col={ws.max_column}")

# ============================================================
# 9. Buscar en shared strings
# ============================================================
print("\n" + "=" * 80)
print("9. PALABRAS CLAVE EN SHARED STRINGS")
print("=" * 80)
with open('temp_xl/xl/sharedStrings.xml', 'r', encoding='utf-8') as f:
    ss = f.read()
keywords = ['tasa', 'imprevisto', 'porcentaje', '%', 'recuperación', 'honorario', 'formula']
for kw in keywords:
    if kw.lower() in ss.lower():
        print(f"  '{kw}' encontrado en shared strings")

# ============================================================
# 10. Extraer contenido completo de sheet5 (TASA DE IMPREVISTOS) desde XML
# ============================================================
print("\n" + "=" * 80)
print("10. XML BRUTO DE 'TASA DE IMPREVISTOS' (sheet5.xml)")
print("=" * 80)
with open('temp_xl/xl/worksheets/sheet5.xml', 'r', encoding='utf-8') as f:
    sheet5_xml = f.read()
print(sheet5_xml[:3000])
print("...")
print(sheet5_xml[-2000:])

print("\n" + "=" * 80)
print("ANÁLISIS COMPLETADO")
print("=" * 80)
